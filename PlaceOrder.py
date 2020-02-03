from ibapi.client import EClient
from ibapi.wrapper import EWrapper
from ibapi.contract import Contract
from ibapi.order import Order
from threading import Timer
import openpyxl as Workbook
import getpass

class TestApp(EClient,EWrapper):
    def __init__(self):
        EClient.__init__(self,self)
    def error(self, reqId, errorCode, errorString):
        print(reqId,"error",errorString)
    def nextValidId(self,orderId):
        self.nextOrderId=orderId
        self.start()
    def orderStatus(self, orderId, status, filled,
                    remaining, avgFillPrice, permId,
                    parentId:int, lastFillPrice, clientId,
                    whyHeld, mktCapPrice):
        print("OrderStatus. Id :", orderId,"Status",status)
    def openOrder(self, orderId, contract, order,
                  orderState):
        print("OpenOrder. ID:", orderId, contract.symbol,order.totalQuantity)
    def execDetails(self, reqId, contract, execution):
        print("ExecDetails:",reqId, contract.symbol,execution.shares)
    def start(self):
        contract=Contract()
        contract.symbol = "TGP"
        contract.secType = "STK"
        contract.exchange = "SMART"
        contract.currency = "USD"
        contract.primaryExchange = "NASDAQ"

        wb = Workbook.load_workbook('D:\\TWS API\\source\\pythonclient\\ibapi\\Accounting Info TGP.xlsx')
        range = wb['Accounting Info']





        wb_1 = Workbook.load_workbook('D:\\TWS API\\source\\pythonclient\\ibapi\\StockPrice TGP.xlsx')
        range_1 = wb_1['stock yield TGP']



        wb_2 = Workbook.load_workbook('D:\\TWS API\\source\\pythonclient\\ibapi\\Dividend Info TGP.xlsx')
        range_2 = wb_2['Dividend Info']

        wb_3 = Workbook.load_workbook('D:\\TWS API\\source\\pythonclient\\ibapi\\StockPrice TGP PRa.xlsx')
        range_3 = wb_3['stock yield TGP PRa']

        wb_4= Workbook.load_workbook('D:\\TWS API\\source\\pythonclient\\ibapi\\YTC TGP.xlsx')
        range_4 = wb_4['YTC TGP']
        TGP_Dividend_PRa=0.5
        TGP_Dividend_PRb=0.4
        Outstanding_share=1000
        condition_1=TGP_Dividend_PRa>TGP_Dividend_PRb
        condition_2=float(str(range_4['A1'].value))>0.7
        condition_3=float(str(range_3['A1'].value))<float(str(range_1['A1'].value))
        condition_4=float(str(range['A1'].value))-float(str(range_2['B1'].value))*Outstanding_share





        if condition_1:
            if condition_2 and condition_3 and condition_4:
                order= Order()
                order.action="BUY"
                order.orderType="LMT"
                order.totalQuantity = 400
                order.lmtPrice=210

        if not(condition_1 and condition_2 and condition_3 and condition_4):
            print('doesnt meet criteria cant place order')
        else:
            self.placeOrder(self.nextOrderId, contract,order)
    def stop(self):
        self.done
        self.disconnect()

def main():
    app=TestApp()
    app.nextOrderId=0
    app.connect("127.0.0.1",7497,9)
    Timer(3,app.stop).start()
    app.run()


if __name__=='__main__':
    main()


