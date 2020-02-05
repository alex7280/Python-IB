from ibapi.client import EClient
from ibapi.wrapper import EWrapper
from ibapi.contract import Contract
from threading import Timer
import numpy as np
import openpyxl as Workbook
import getpass

class TestApp(EClient,EWrapper):

    def __init__(self):
        EClient.__init__(self,self)
        self.barsList=[]
    def error(self, reqId, errorCode, errorString):
        print(reqId,"error",errorString)
    def nextValidId(self,reqId):
        self.start()



    def weightedAvg(self, barsList):
        self.weights = np.repeat(1.0, len(barsList)) / len(barsList)
        self.smas = np.convolve(barsList, self.weights, 'valid')
        return self.smas
    def historicalData(self, reqId, bar):
        avg=bar.close
        self.barsList.append(avg)
        if self.weightedAvg(self.barsList)>1000:
            print("error")
        else:
            print(self.weightedAvg(self.barsList))
            wb = Workbook.Workbook()

            ws1 = wb.active
            ws1.title = "stock yield"
            a=self.weightedAvg(self.barsList).tolist()
            
            ws1.append(a)
            i=0
            wb.save('C:\\Users\\' + getpass.getuser() + '\\Desktop\\'+str(i)+'.xlsx')
            i+=1
        print(bar.date)


    def start(self):
        contract = Contract()
        wb = Workbook.load_workbook('StockTickers.xlsx')
        sheet_ranges=wb['Sheet1']


        empty_list=[]
        for row in range(1, sheet_ranges.max_row + 1):
            for column in "A":
                cell_name = "{}{}".format(column, row)
                empty_list.append(sheet_ranges[cell_name].value)
        contract.secType = "STK"
        contract.exchange = "SMART"
        contract.currency = "USD"
        contract.primaryExchange = "NASDAQ"
        for i in empty_list:
            contract.symbol=str(i)
            print(i)
            self.reqHistoricalData(1, contract, "", "26 W", "1 day", "MIDPOINT", 0, 1, False, [])
            self.tickerId=+1

    def stop(self):
        self.done=True
        self.disconnect()



def main():

    app = TestApp()
    app.nextOrderId=0
    app.connect("127.0.0.1", 7497, 0)
    Timer(3, app.stop).start()


    app.run()




if __name__=="__main__":
    main()