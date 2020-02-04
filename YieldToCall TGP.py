import openpyxl as Workbook
from datetime import datetime
import getpass

def main():
    wb = Workbook.load_workbook('D:\\TWS API\\source\\pythonclient\\ibapi\\Dividend Info TGP.xlsx')
    range=wb['Dividend Info']

    date=str(range['A1'].value)
    annual_interest = float(str(range['B1'].value))

    wb_1= Workbook.load_workbook('D:\\TWS API\\source\\pythonclient\\ibapi\\StockPrice TGP.xlsx')
    range_1=wb_1['stock yield TGP']
    price=float(str(range_1['A1'].value))



    now_datetime=datetime.now()
    date_year=date[:4]
    date_month=date[4:6]
    date_day=date[6:8]
    total_next_dividend=float(date_year)*365+float(date_month)*30+float(date_day)


    now_year=str(now_datetime)[:4]
    now_month=str(now_datetime)[5:7]
    now_day=str(now_datetime)[8:10]
    now_total=float(now_year)*365+float(now_month)*30+float(now_day)
    time_difference=(total_next_dividend-now_total)/365

    Call_Price=float(input('Please enter call price'))
    YTC=(annual_interest+(price-Call_Price)/2)/((price+Call_Price)/2)

    wb = Workbook.Workbook()

    ws1 = wb.active
    ws1.title = "YTC TGP"
    a = [YTC]

    ws1.append(a)
    wb.save('D:\\TWS API\\source\\pythonclient\\ibapi\\YTC TGP.xlsx')





if __name__=='__main__':
    main()
